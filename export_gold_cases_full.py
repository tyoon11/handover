#!/usr/bin/env python3
"""
export_gold_cases_full.py — gold 케이스별 검토 엑셀 (원본 EMR '전체', 캡 없음)

왜 필요한가
  report_v3 의 리포트 HTML 은 EMR 을 '표시용 2600자'로 캡(_EMR_DISPLAY_CAP)해서,
  긴 케이스는 뒷부분이 잘린다. 게다가 그 표시 절단은 '뒤(최신)'를 자르지만 모델 실제
  입력은 '앞(오래된 쪽)'을 좌측절단한다(prompt_utils). → 검토용으로 '완전한 원본 EMR'이
  필요하면 리포트가 아니라 소스 pkl 에서 직접 build_emr_text 로 다시 뽑아야 한다.

무엇을 뽑나 (케이스=수술ID 시트마다)
  ① Input · EMR (비식별, 캡 없음)  — 4대 기록지: Pre-anesthetic Summary / Premedication
                                     / Anesthetic TOTALS / Anesthetic Record
  ② 전문의 GT (정답)                — KHS gold (인계요약지_gold_sampled_..._KHS.xlsx)
  ③ 바이탈 요약                     — vital_summary_map (규칙기반, 연령별 임계값)
  ④ 모델별 v3 추론결과 (21변형)     — INFER_OUT 생성문 + EVAL_OUT 점수(comp/cov/faith/brev/gate)
                                     + 다린(기존) 출력 병기(llama·qwen35)
  + 별도 시트: Threshold(연령별 바이탈 임계값), 검토(전체 요약; report 병기와 동일 데이터)

의존
  pipeline_v3 (config_v3, data_splits, prompt_utils, eval_v3.checklist, report_v3),
  pandas, openpyxl. torch 불필요.

실행 (v3 repo 루트, gold 데이터 있는 서버)
  HANDOVER_RUN_ID=v3_20260709 python export_gold_cases_full.py \
      --darin_root ~/workspace/data/HANDOVER_인계용_다린/data/inferenced_v3sids \
      --out output_케이스별_full.xlsx
  # --darin_root 생략 시 config DARIN_INFER_OUT 사용. --no_darin 로 다린 병기 끄기.
  # --cap N 으로 EMR 표시 길이 상한(기본 0 = 무제한).
"""
import argparse
import json
import re
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from pipeline_v3.config_v3 import (
    DARIN_INFER_OUT, EVAL_OUT, GOLD_KHS_XLSX, GOLD_PKL, INFER_OUT, VITAL_MAP_PKL,
)
from pipeline_v3.data_splits import load_splits
from pipeline_v3.prompt_utils import build_emr_text, get_sid
from pipeline_v3.report_v3 import _load_darin_outputs, _scrub_phi

MODELS = ["llama", "qwen35", "gemma4"]
VORDER = ["raw", "sft_1ep", "sft_3ep", "rlaif_dpo", "rlaif_simpo",
          "sft_1ep_dpo", "sft_3ep_dpo"]
EMR_SECTIONS = [
    ("Pre-anesthetic Patient Conditions Summary", "마취전 환자상태 요약"),
    ("Preoperative Preparations and Premedication", "수술전 준비사항·Premedication"),
    ("Anesthetic TOTALS", "마취기록 TOTALS"),
    ("Anesthetic Record", "마취기록 본문"),
]


def split_emr_sections(emr_text: str):
    """build_emr_text 출력(4개 '- <라벨>\\n내용' 블록)을 (라벨, 내용) 리스트로 분리."""
    labels = [lab for lab, _ in EMR_SECTIONS]
    pat = re.compile(r"- (" + "|".join(re.escape(l) for l in labels) + r")\n")
    parts, pos, cur_lab = [], 0, None
    out = []
    for m in pat.finditer(emr_text):
        if cur_lab is not None:
            out.append((cur_lab, emr_text[pos:m.start()].strip()))
        cur_lab = m.group(1)
        pos = m.end()
    if cur_lab is not None:
        out.append((cur_lab, emr_text[pos:].strip()))
    return out or [("EMR", emr_text.strip())]


def load_sources(split, cap, scrub):
    """{sid: {'emr_sections':[(lab,txt)], 'vital', 'gt'}} — 원본에서 직접(캡 옵션)."""
    splits = load_splits(need=(split,))
    df = splits[split]
    with open(VITAL_MAP_PKL, "rb") as f:
        import pickle
        vital_map = pickle.load(f)

    gt_map = {}
    try:
        from pipeline_v3.eval_v3 import checklist as CK
        gold_df = pd.read_pickle(GOLD_PKL)
        gold_refs, _ = CK.load_khs_gold(GOLD_KHS_XLSX, gold_df)
        for i in range(len(gold_df)):
            s = get_sid(gold_df.iloc[i])
            if s != -1 and i in gold_refs:
                gt_map[s] = gold_refs[i]
    except Exception as e:
        print(f"[warn] GT 로드 실패 → GT 생략 ({type(e).__name__}: {e})")

    out = {}
    for i in range(len(df)):
        row = df.iloc[i]
        s = get_sid(row)
        if s == -1:
            continue
        emr = build_emr_text(row)
        if scrub:
            emr = _scrub_phi(emr)
        if cap and len(emr) > cap:
            emr = emr[:cap] + " …(이하 생략)"
        secs = split_emr_sections(emr)
        vit = vital_map.get(s, "")
        if scrub:
            vit = _scrub_phi(str(vit))
        gt = gt_map.get(s, "")
        out[s] = dict(emr_sections=secs, vital=str(vit),
                      gt=(_scrub_phi(str(gt)) if scrub else str(gt)) or "")
    return out, list(out.keys())


def load_model_outputs(split):
    """{tag: {sid: {'gen','comp','cov','faith','brev','gate','excluded'}}} —
    생성문은 INFER_OUT, 점수는 EVAL_OUT 에서."""
    out = {}
    for tag_dir in sorted(INFER_OUT.iterdir()) if INFER_OUT.exists() else []:
        inf = tag_dir / f"{split}_results.jsonl"
        if not inf.exists():
            continue
        tag = tag_dir.name
        gen = {}
        for line in inf.read_text(encoding="utf-8").splitlines():
            if not line.strip():
                continue
            r = json.loads(line)
            gen[r.get("sid")] = r.get("generated", "")
        recs = {}
        sc = EVAL_OUT / tag / f"{split}_results_scores_v3.jsonl"
        scores = {}
        if sc.exists():
            for line in sc.read_text(encoding="utf-8").splitlines():
                if not line.strip():
                    continue
                r = json.loads(line)
                scores[r.get("sid")] = r.get("official") or {}
        for sid, g in gen.items():
            off = scores.get(sid, {})
            recs[sid] = dict(gen=g, comp=off.get("composite"), cov=off.get("coverage"),
                             faith=off.get("faithfulness"), brev=off.get("brevity"),
                             gate=off.get("gate"), excluded=off.get("excluded"))
        out[tag] = recs
    return out


# ── 스타일 ──────────────────────────────────────────────────────────────────
HDR = PatternFill("solid", fgColor="305496"); GTF = PatternFill("solid", fgColor="E2EFDA")
INF = PatternFill("solid", fgColor="FFF2CC"); V3F = PatternFill("solid", fgColor="DDEBF7")
DARF = PatternFill("solid", fgColor="F2F2F2"); CMTF = PatternFill("solid", fgColor="FCE4D6")
ALT = PatternFill("solid", fgColor="F2F6FC")
WB_ = Font(bold=True, color="FFFFFF"); BOLD = Font(bold=True)
THIN = Side("thin", color="C7C7C7"); PATB = Side("medium", color="4472C4")
BORDER = Border(THIN, THIN, THIN, THIN)
WRAP = Alignment(wrap_text=True, vertical="top")
CTR = Alignment(horizontal="center", vertical="center", wrap_text=True)
LTOP = Alignment(horizontal="left", vertical="top", wrap_text=True)


def _b(ws, r, c1, c2, fill=None, font=None, al=None):
    for c in range(c1, c2 + 1):
        cell = ws.cell(r, c); cell.border = BORDER
        if fill: cell.fill = fill
        if font: cell.font = font
        if al: cell.alignment = al


# 나이 대표값 — 임계값 표를 vital_thresholds에서 직접 뽑아 쓰기 위한 샘플 연령.
# 표를 손으로 적어두면 임계값 개정 때 검토용 산출물만 옛 값으로 남는 사고가 난다.
_TH_AGES = [("신생아\n(14일)", 14 / 365.0), ("영아\n(6개월)", 0.5),
            ("유아\n(2세)", 2.0), ("학령전\n(4세)", 4.0),
            ("학령기\n(7세)", 7.0), ("학령기\n(10세)", 10.0),
            ("청소년\n(14세)", 14.0)]


def _th_rows():
    """vital_thresholds 실제 함수로 연령군별 임계값 행 생성."""
    import utils.vital_thresholds as T
    ages = [a for _, a in _TH_AGES]
    hr_n = [T.hr_normal_range(a) for a in ages]
    hr_c = [T.hr_critical_range(a) for a in ages]
    return [
        ("HR 정상범위 (bpm)", *[f"{lo:.0f}–{hi:.0f}" for lo, hi in hr_n]),
        ("HR [유의]서맥 < / [유의]빈맥 > (bpm)", *[f"{b:.0f} / {t:.0f}" for b, t in hr_c]),
        ("SBP [유의]저혈압 < (mmHg)", *[f"{T.sbp_hypotension(a):.0f}" for a in ages]),
        ("SBP 고혈압 > (mmHg)", *[f"{T.sbp_hypertension(a):.0f}" for a in ages]),
        ("MBP [유의]저혈압 < (mmHg)", *[f"{T.map_hypotension(a):.0f}" for a in ages]),
        ("DBP 고 > (mmHg)", *[f"{T.dbp_hypertension(a):.0f}" for a in ages]),
        ("QTc 정상상한 > (ms)", *[f"{T.qtc_upper_normal(a):.0f}" for a in ages]),
        ("EBV (mL/kg)", *[f"{T.estimated_blood_volume_ml_per_kg(a):.0f}" for a in ages]),
    ]


def build_threshold_sheet(ws):
    import utils.vital_thresholds as T
    rows = _th_rows()
    ws.sheet_view.showGridLines = False
    ws["A1"] = ("소아 바이탈 임계값 (Threshold) — Smith's Anesthesia 2021 / "
                "Miller's Anesthesia 2024 근거")
    ws["A1"].font = Font(bold=True, size=14, color="1F3864"); ws.merge_cells("A1:G1")
    ws.row_dimensions[1].height = 26
    ws.cell(3, 1, "■ 연령군별 임계값 (나이 = 수술당시나이) · [유의] = 임상적으로 유의") \
        .font = Font(bold=True, size=11, color="1F3864")
    for j, a in enumerate(["항목"] + [lab for lab, _ in _TH_AGES], 1):
        ws.cell(4, j, a)
    _b(ws, 4, 1, 7, HDR, WB_, CTR); ws.row_dimensions[4].height = 36
    for i, row in enumerate(rows):
        r = 5 + i
        for j, v in enumerate(row, 1): ws.cell(r, j, v)
        _b(ws, r, 1, 7, ALT if i % 2 else None, None, CTR)
        ws.cell(r, 1).font = BOLD
        ws.cell(r, 1).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.row_dimensions[r].height = 24

    r0 = 5 + len(rows) + 1
    ws.cell(r0, 1, "■ 고정 임계값 및 판정 제외 항목").font = Font(bold=True, size=11, color="1F3864")
    ws.cell(r0 + 1, 1, "항목"); ws.cell(r0 + 1, 2, "기준"); ws.cell(r0 + 1, 7, "출처")
    ws.merge_cells(start_row=r0 + 1, start_column=2, end_row=r0 + 1, end_column=6)
    _b(ws, r0 + 1, 1, 7, HDR, WB_, CTR); ws.row_dimensions[r0 + 1].height = 22
    t2 = [("SpO2", f"[유의]<{T.SPO2_CRIT:.0f}%    ·    목표미달 {T.SPO2_CRIT:.0f}–"
                   f"{T.SPO2_TARGET_LOW - 1:.0f}% (목표 94–99%)", "Smith Ch.57"),
          ("체온 (T1)", f"[유의]<{T.TEMP_SAFE_LOW} ℃  ·  저체온 <{T.TEMP_HYPOTHERMIA} ℃  ·  "
                        f"안전범위 초과 >{T.TEMP_SAFE_HIGH} ℃  ·  [유의]발열 >{T.TEMP_FEVER} ℃",
           "Smith Ch.21 / Ch.7"),
          ("QTc 연장", f"[유의]>{T.QTC_PROLONGED:.0f} ms (정상상한은 위 표)", "Miller / Smith Ch.5"),
          ("UO 핍뇨", f"[유의]< {T.UO_OLIGURIA} mL/kg/hr (실제 기록 경과시간 기준)", "Miller Ch.24"),
          ("EBL", f"[유의]유의 >{T.EBL_SIGNIFICANT_PCT:.0f}% EBV  ·  "
                  f"[유의]대량 >{T.EBL_MASSIVE_PCT:.0f}% EBV (체중 있을 때만)", "Smith Ch.18 / Table 21.6"),
          ("DBP 하한 / Ppeak", "판정하지 않음 — 교과서에 소아 기준 없음 (수치 요약만)", "—")]
    for i, (item, crit, src) in enumerate(t2):
        r = r0 + 2 + i
        ws.cell(r, 1, item); ws.cell(r, 2, crit); ws.cell(r, 7, src)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
        _b(ws, r, 1, 7, ALT if i % 2 else None, None, None)
        ws.cell(r, 1).font = BOLD; ws.cell(r, 1).alignment = LTOP
        ws.cell(r, 2).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.cell(r, 7).alignment = CTR; ws.row_dimensions[r].height = 24
    rn = r0 + 2 + len(t2) + 1
    ws.cell(rn, 1, "※ 나이 = 수술당시나이(세). 바이탈 값 0은 센서 미연결로 제외. "
                   "각 케이스 '바이탈 요약'의 이벤트는 위 임계값으로 판정됨. "
                   "표·페이지 단위 전거는 docs/THRESHOLDS.md 참고.")
    ws.merge_cells(start_row=rn, start_column=1, end_row=rn + 1, end_column=7)
    ws.cell(rn, 1).font = Font(italic=True, color="7F7F7F")
    ws.cell(rn, 1).alignment = Alignment(wrap_text=True, vertical="top")
    for col, w in zip("ABCDEFG", [26, 16, 16, 16, 16, 16, 18]):
        ws.column_dimensions[col].width = w


def build_case_sheet(ws, sid, src, mo, darin):
    ws.sheet_view.showGridLines = False
    ws.cell(1, 1, f"수술 ID {sid}   ·   ① Input EMR(원본 전체) → ② 전문의 GT · ③ 바이탈 → ④ v3 21변형(+다린)")
    ws.cell(1, 1).font = Font(bold=True, size=13, color="1F3864"); ws.merge_cells("A1:I1")
    ws.row_dimensions[1].height = 24
    r = 3
    # ① EMR (섹션별, 접기 가능)
    ws.cell(r, 1, "① Input · EMR (비식별, 원본 전체 — 캡 없음)").font = Font(bold=True, size=11, color="1F3864")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9); r += 1
    label_map = dict(EMR_SECTIONS)
    for lab, txt in src["emr_sections"]:
        ko = label_map.get(lab, lab)
        ws.cell(r, 1, f"◦ {lab}  ({ko})"); ws.cell(r, 1).font = BOLD
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
        ws.cell(r, 1).fill = INF; _b(ws, r, 1, 9); r += 1
        ws.cell(r, 1, txt or "(없음)"); ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
        ws.cell(r, 1).alignment = LTOP; _b(ws, r, 1, 9)
        ws.row_dimensions[r].height = min(320, max(30, (len(txt) // 90 + txt.count("\n") + 1) * 15))
        ws.row_dimensions[r].outline_level = 1   # 접기(그룹)
        r += 1
    r += 1
    # ② GT / ③ 바이탈 (나란히)
    ws.cell(r, 1, "② 전문의 GT (정답)"); ws.cell(r, 5, "③ 바이탈 요약 (규칙기반)")
    for c in (1, 5): ws.cell(r, c).font = BOLD; ws.cell(r, c).fill = HDR; ws.cell(r, c).font = WB_
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=9)
    _b(ws, r, 1, 9, HDR, WB_, CTR); r += 1
    ws.cell(r, 1, src["gt"] or "(없음)"); ws.cell(r, 5, src["vital"] or "(없음)")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=9)
    ws.cell(r, 1).fill = GTF
    for c in (1, 5): ws.cell(r, c).alignment = LTOP
    _b(ws, r, 1, 9)
    ws.row_dimensions[r].height = max(60, (max(len(src["gt"]), len(src["vital"])) // 40 + 2) * 15)
    r += 2
    # ④ 모델 결과
    ws.cell(r, 1, "④ 모델별 v3 추론결과 (3 family × 7 variant = 21) + 다린 병기").font = Font(bold=True, size=11, color="1F3864")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9); r += 1
    H = ["모델", "변형", "v3 추론결과", "다린(기존)", "comp", "cov", "faith", "brev", "게이트"]
    for j, x in enumerate(H, 1): ws.cell(r, j, x)
    _b(ws, r, 1, 9, HDR, WB_, CTR); ws.row_dimensions[r].height = 26; r += 1
    num = lambda x: (round(float(x), 3) if isinstance(x, (int, float)) else None)
    for mi, mdl in enumerate(MODELS):
        first = True
        for v in VORDER:
            tag = f"{mdl}_{v}"; rec = (mo.get(tag) or {}).get(sid)
            if rec is None:
                continue
            d = (darin.get(tag) or {}).get(sid)
            ws.cell(r, 1, mdl if first else ""); ws.cell(r, 2, v); first = False
            if rec.get("excluded"):
                ws.cell(r, 3, "(제외)")
            else:
                ws.cell(r, 3, rec.get("gen") or "")
                ws.cell(r, 4, d if d else ("— (다린 없음)" if mdl == "gemma4" else ""))
                for cc, key in ((5, "comp"), (6, "cov"), (7, "faith"), (8, "brev")):
                    ws.cell(r, cc, num(rec.get(key)))
                if rec.get("gate"): ws.cell(r, 9, rec.get("gate"))
            _b(ws, r, 1, 9, None, None, WRAP)
            ws.cell(r, 1).font = BOLD; ws.cell(r, 3).fill = V3F; ws.cell(r, 4).fill = DARF
            for cc in (5, 6, 7, 8): ws.cell(r, cc).alignment = Alignment(horizontal="right", vertical="top")
            if rec.get("gate"): ws.cell(r, 9).font = Font(bold=True, color="C00000")
            g = rec.get("gen") or ""
            ws.row_dimensions[r].height = min(150, max(30, (len(g) // 46 + 1) * 15))
            last = (v == VORDER[-1] and mi < len(MODELS) - 1)
            if last:
                for c in range(1, 10):
                    cur = ws.cell(r, c).border
                    ws.cell(r, c).border = Border(cur.left, cur.right, cur.top, PATB)
            r += 1
    for col, w in zip("ABCDEFGHI", [11, 13, 60, 34, 7, 6, 7, 6, 13]): ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"


# ── Markdown 출력 (복붙 친화 — EMR/GT/바이탈은 코드블록, 모델출력은 인용문) ──────
def build_threshold_md() -> str:
    """xlsx Threshold 시트와 같은 값을 vital_thresholds에서 뽑아 markdown으로."""
    import utils.vital_thresholds as T
    labs = [lab.replace("\n", " ") for lab, _ in _TH_AGES]
    head = ("| 항목 | " + " | ".join(labs) + " |\n"
            + "|---" * (len(labs) + 1) + "|\n")
    body = "".join(
        "| " + " | ".join(str(c).replace("<", "&lt;").replace(">", "&gt;")
                          for c in row) + " |\n"
        for row in _th_rows())
    return f"""## 바이탈 임계값 (Threshold) — Smith's Anesthesia 2021 / Miller's Anesthesia 2024 근거

[유의] = 소생·개입 기준 초과(임상적으로 유의) · 표시 없음 = 연령별 정상 참조범위 이탈

### 연령별 (나이 = 수술당시나이)

임계값은 **연속 나이로 계산**된다. 아래는 괄호 안 **대표 나이**에서의 값이며, 실제 판정에는
그 케이스의 나이가 그대로 쓰인다 (HR 정상범위는 Smith Table 18.1의 9개 원구간).

{head}{body}
### 고정 임계값 및 판정 제외 항목

| 항목 | 기준 | 출처 |
|---|---|---|
| SpO2 | [유의]&lt;{T.SPO2_CRIT:.0f}% · 목표미달 {T.SPO2_CRIT:.0f}–{T.SPO2_TARGET_LOW - 1:.0f}% (목표 94–99%) | Smith Ch.57 |
| 체온 T1 | [유의]&lt;{T.TEMP_SAFE_LOW}℃ · 저체온 &lt;{T.TEMP_HYPOTHERMIA}℃ · 안전범위 초과 &gt;{T.TEMP_SAFE_HIGH}℃ · [유의]발열 &gt;{T.TEMP_FEVER}℃ | Smith Ch.21 / Ch.7 |
| QTc 연장 | [유의]&gt;{T.QTC_PROLONGED:.0f}ms (정상상한은 위 표) | Miller / Smith Ch.5 |
| UO 핍뇨 | [유의]&lt; {T.UO_OLIGURIA} mL/kg/hr (실제 기록 경과시간 기준) | Miller Ch.24 |
| EBL | [유의]유의 &gt;{T.EBL_SIGNIFICANT_PCT:.0f}% EBV · [유의]대량 &gt;{T.EBL_MASSIVE_PCT:.0f}% EBV (체중 있을 때만) | Smith Ch.18 / Table 21.6 |
| DBP 하한 / Ppeak | 판정하지 않음 — 교과서에 소아 기준 없음 (수치 요약만) | — |

※ 나이 = 수술당시나이(세). 바이탈 값 0은 센서 미연결로 제외. 각 케이스 '바이탈 요약'의 이벤트는 위 임계값으로 판정됨.
표·페이지 단위 전거는 `docs/THRESHOLDS.md`.
"""


def _fmt(x):
    return f"{float(x):.3f}" if isinstance(x, (int, float)) else "—"


def build_markdown(src, sids, mo, darin):
    L = ["# Gold 검토 — Input · 전문의 GT · v3 추론 · 다린\n",
         "> ⚠ **PHI 포함** — 외부 공유·커밋 금지. gold checklist 미검수면 점수는 잠정치.",
         "> composite = 0.5·coverage + 0.3·faithfulness + 0.2·brevity  ·  "
         "안전게이트 missed_abnormal = 이상소견을 '특이사항 없음'으로 뭉갬 → composite 0\n",
         build_threshold_md()]
    for sid in sids:
        s = src[sid]
        L.append(f"\n---\n\n## 수술 ID {sid}\n")
        L.append("### ① Input · EMR (비식별, 원본 전체)\n")
        for lab, txt in s["emr_sections"]:
            L.append(f"**{lab}**\n\n```\n{txt or '(없음)'}\n```\n")
        L.append("### ② 전문의 GT (정답)\n\n```\n" + (s["gt"] or "(없음)") + "\n```\n")
        L.append("### ③ 바이탈 요약 (규칙기반, 연령별 임계값)\n\n```\n" + (s["vital"] or "(없음)") + "\n```\n")
        L.append("### ④ 모델별 v3 추론결과 (+ 다린 병기)\n")
        for mdl in MODELS:
            if not any((mo.get(f"{mdl}_{v}") or {}).get(sid) for v in VORDER):
                continue
            L.append(f"\n#### {mdl}\n")
            for v in VORDER:
                rec = (mo.get(f"{mdl}_{v}") or {}).get(sid)
                if rec is None:
                    continue
                if rec.get("excluded"):
                    L.append(f"**{v}** — (제외)\n")
                    continue
                sc = f"comp {_fmt(rec.get('comp'))} · cov {_fmt(rec.get('cov'))} · " \
                     f"faith {_fmt(rec.get('faith'))} · brev {_fmt(rec.get('brev'))}"
                gate = f"  ·  ⚠ **{rec['gate']}**" if rec.get("gate") else ""
                L.append(f"**{v}**  ({sc}){gate}")
                out = (rec.get("gen") or "").strip() or "(빈 출력)"
                L.append("> " + out.replace("\n", "\n> "))
                d = (darin.get(f"{mdl}_{v}") or {}).get(sid)
                if d:
                    L.append(f"\n_다린:_ {d.strip()}")
                elif mdl == "gemma4":
                    L.append("\n_다린: — (없음)_")
                L.append("")
    return "\n".join(L)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--split", default="gold", choices=["gold", "dev"])
    ap.add_argument("--format", default="md", choices=["md", "xlsx"])
    ap.add_argument("--out", default=None, help="기본: output_케이스별_full.{md|xlsx}")
    ap.add_argument("--darin_root", default=str(DARIN_INFER_OUT))
    ap.add_argument("--no_darin", action="store_true")
    ap.add_argument("--cap", type=int, default=0, help="EMR 표시 길이 상한(0=무제한)")
    ap.add_argument("--no_scrub", action="store_true", help="식별자 스크럽 끄기(원본 그대로)")
    args = ap.parse_args()
    out = args.out or f"output_케이스별_full.{args.format}"

    print(f"[export] split={args.split} format={args.format} "
          f"cap={args.cap or '무제한'} scrub={not args.no_scrub}")
    src, sids = load_sources(args.split, args.cap, not args.no_scrub)
    mo = load_model_outputs(args.split)
    darin = {} if args.no_darin else _load_darin_outputs(args.darin_root)
    sids_sorted = sorted(sids, key=lambda x: int(x) if str(x).isdigit() else 0)
    print(f"[export] 케이스 {len(sids_sorted)} · v3 변형 {len(mo)} · 다린 변형 {len(darin)}")

    if args.format == "md":
        Path(out).write_text(build_markdown(src, sids_sorted, mo, darin), encoding="utf-8")
    else:
        wb = Workbook()
        build_threshold_sheet(wb.active); wb.active.title = "Threshold"
        for sid in sids_sorted:
            build_case_sheet(wb.create_sheet(str(sid)), sid, src[sid], mo, darin)
        wb.save(out)
    print(f"[export] 저장: {out}  (케이스 {len(sids_sorted)})")
    print("  ⚠ PHI 포함 — 외부 공유·커밋 금지. gold checklist 미검수면 점수는 잠정치.")


if __name__ == "__main__":
    main()
