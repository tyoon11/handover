#!/usr/bin/env python3
"""
import_khs_review.py — 교수님 검수본(output_케이스별_*_reviewed.xlsx) → 구조화 추출

export_gold_cases_full.py 가 만든 케이스별 xlsx에 교수님이 직접 적어 넣은
'②-1 교수님 검토 의견'(GT 수정·보완 사항 / 기타 의견)을 뽑아 정리한다.

산출물 (전부 PHI — data/ 밑, gitignored):
  1) khs_review_<stamp>.json          : 케이스별 원문 + 자동 분류
  2) khs_gt_merge_worksheet.xlsx      : final_gt 확정용 검토 시트
  3) (확정 후 수기 생성) khs_gold_override.json : {sid: final_gt}

병합 방침 (연구책임자 확정, 2026-08-14)
--------------------------------------
  1. 'GT 수정·보완 사항'이 비어 있으면  → 기존 전문의 GT 를 그대로 쓴다.
  2. 기존 GT를 **보완**하는 내용이면    → 기존 GT + 보완 내용을 합쳐 GT로 삼는다.
  3. 통째로 다시 쓴 문안이면            → 그것으로 교체한다.
  4. 기존 GT를 **부정**하는 문안이면    → 검토 의견을 우선해 교체한다.

2와 4의 구분은 문장만으로 기계 판정이 안 된다(같은 "수술 중 특이사항 없었음"이라도,
기존 GT가 기저질환·기도 정보면 보완이고 기존 GT가 수술 중 처치 기록이면 부정이다).
그래서 이 스크립트는 **기본값으로 병합**하되, 합친 결과에 실질 소견과 '특이사항 없음'
주장이 동시에 들어가면 `conflict_check` 로 표시해 사람이 보게 한다. 실제 확정값은
워크시트('최종 GT' 열)에 있고, 그 파일이 단일 출처다 (repo에는 sid를 두지 않는다).

사용:
  python scripts/import_khs_review.py output_케이스별_khs_reviewed.xlsx
  python scripts/import_khs_review.py <xlsx> --show      # 최종 GT 전문 출력
"""

import argparse
import json
import re
import sys
from datetime import datetime
from pathlib import Path

import openpyxl

GT_ANCHOR = "②  전문의 GT"
REVIEW_HDR = "GT 수정"

# 앵커 행 기준 오프셋 (export_gold_cases_full.build_case_sheet 구조)
OFF_GT_ROW, COL_GT, COL_VITAL = 1, 1, 5
OFF_REVIEW_HDR, OFF_REVIEW_ROW, COL_NEW_GT, COL_ETC = 3, 4, 1, 7

# 기존 GT 본문을 텍스트로 가리키는 표현 → 단독으로 쓰면 지시대상이 사라지므로 반드시 병합
_REFERENCES_PRIOR = re.compile(r"(위의?\s*내용|상기|앞의?\s*내용|이\s*외|추가\s*$)")
# '이상/필요 없음' 주장 — 기존 GT를 부정할 수 있는 표현
_NO_ISSUE = re.compile(
    r"(특이\s*사항\s*(없|없었|없음)|이상\s*(소견\s*)?없|필요\s*없|문제\s*없)")
# 실질 임상 소견·처치로 볼 만한 신호 (병합 결과가 자기모순인지 확인용)
_SUBSTANTIVE = re.compile(
    r"(투여|투약|수혈|삽관|재삽관|처치|교체|suction|흡인|wheez|crackle|desat|"
    r"탈포화|저혈압|서맥|부정맥|출혈|경련|VT|arrest|심정지|puff|ventolin|"
    r"lipoma|compression|병력|증후군)", re.I)

# 통째 재작성으로 볼 최소 길이 — 이보다 길면 보완이 아니라 대체로 본다
FULL_REWRITE_MIN = 100


def _clean(v) -> str:
    """셀 값 정규화.

    Excel 왕복에서 붙는 바깥 큰따옴표를 벗긴다. 검수본에는 여는 따옴표만 남고
    닫는 쪽이 없는 셀도 있어(수기 편집 중 잘림) 짝이 안 맞는 경우까지 처리한다.
    """
    if v is None:
        return ""
    s = str(v).strip()
    if s.lower() in ("nan", "none", "-", "(없음)"):
        return ""
    if len(s) >= 2 and s[0] == '"' and s[-1] == '"':
        s = s[1:-1].strip()
    elif s.count('"') % 2 == 1:          # 짝 없는 따옴표 — 바깥쪽만 제거
        if s.startswith('"'):
            s = s[1:].strip()
        elif s.endswith('"'):
            s = s[:-1].strip()
    return s


def _find_anchor(rows) -> int:
    for i, r in enumerate(rows):
        if r and str(r[0] or "").startswith(GT_ANCHOR):
            return i
    return -1


def _is_no_issue_only(text: str) -> bool:
    """'특이사항 없음' 류 문구를 걷어내면 남는 내용이 거의 없는가.

    키워드 목록을 늘려가며 맞추는 대신 '알맹이가 남는지'로 판정한다.
    """
    body = _NO_ISSUE.sub("", text)
    body = re.sub(r"[\s*#:·\-–—.,()]+", "", body)
    return len(body) < 15


def resolve(old_gt: str, new_gt: str) -> tuple:
    """(최종 GT, 분류) 산출.

    kept            : 검토 의견 없음 → 기존 GT 유지
    empty           : 양쪽 다 비어 gold 없음 (no_gold)
    replaced        : 통째 재작성 문안 → 교체
    merged          : 보완 문안 → 기존 GT + 보완 내용
    merged_conflict : 병합했으나 실질 소견 + '특이사항 없음'이 공존 → 사람 확인
    """
    if not new_gt:
        return (old_gt, "kept") if old_gt else ("", "empty")
    if not old_gt:
        return new_gt, "replaced"
    # 기존 GT 서두를 그대로 다시 쓴 경우 = 통째 재작성
    head = re.sub(r"\W+", "", old_gt)[:12]
    if head and re.sub(r"\W+", "", new_gt).startswith(head):
        return new_gt, "replaced"
    if len(new_gt) >= FULL_REWRITE_MIN and not _REFERENCES_PRIOR.search(new_gt):
        return new_gt, "replaced"
    # 기존 GT가 '특이사항 없음'뿐인데 새 문안에 실질 소견이 있으면 보완이 아니라 대체
    if _is_no_issue_only(old_gt) and not _is_no_issue_only(new_gt):
        return new_gt, "replaced"

    merged = f"{old_gt.rstrip()}\n{new_gt.lstrip()}"
    # "위의 내용 이외에", "~ 이 외", "~ 추가" 는 기존 내용을 명시적으로 남겨두는
    # 표현이므로 '없음'이 함께 있어도 모순이 아니다.
    if _REFERENCES_PRIOR.search(new_gt):
        return merged, "merged"
    if _NO_ISSUE.search(new_gt) and _SUBSTANTIVE.search(old_gt):
        # 기존 GT에 실제 소견·처치가 있는데 '없음/필요없음'을 덧붙이면 자기모순
        return merged, "merged_conflict"
    return merged, "merged"


def extract(xlsx: Path, prefer_review_on_conflict: bool = False) -> list:
    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    out = []
    for name in wb.sheetnames:
        if not name.isdigit():
            continue
        rows = list(wb[name].iter_rows(values_only=True))
        anc = _find_anchor(rows)
        if anc < 0:
            print(f"  ⚠ {name}: GT 앵커 없음 — 건너뜀", file=sys.stderr)
            continue

        def cell(dr, col):
            r = rows[anc + dr] if anc + dr < len(rows) else None
            return _clean(r[col - 1]) if r and len(r) >= col else ""

        hdr = cell(OFF_REVIEW_HDR, 1)
        if REVIEW_HDR not in hdr:
            print(f"  ⚠ {name}: 검토란 헤더 불일치({hdr!r}) — 오프셋 확인 필요",
                  file=sys.stderr)

        old_gt = cell(OFF_GT_ROW, COL_GT)
        new_gt = cell(OFF_REVIEW_ROW, COL_NEW_GT)
        fin, mtype = resolve(old_gt, new_gt)
        if mtype == "merged_conflict" and prefer_review_on_conflict:
            # 방침 4: 기존 GT를 부정하는 문안은 검토 의견을 우선해 교체
            fin, mtype = new_gt, "replaced_conflict"
        out.append({
            "sid": int(name),
            "old_gt": old_gt,
            "new_gt": new_gt,
            "etc": cell(OFF_REVIEW_ROW, COL_ETC),
            "vital_summary": cell(OFF_GT_ROW, COL_VITAL),
            "merge_type": mtype,
            "final_gt": fin,
        })
    return out


def write_worksheet(recs: list, path: Path):
    from openpyxl.styles import Alignment, Font, PatternFill
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "GT 병합 확정"
    hdr = ["수술ID", "분류", "기존 GT", "교수님 수정·보완", "기타 의견",
           "최종 GT (여기 확정)", "확정?"]
    ws.append(hdr)
    fill = PatternFill("solid", fgColor="1F3864")
    warn = PatternFill("solid", fgColor="FFC7CE")
    auto = PatternFill("solid", fgColor="FFF2CC")
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = fill
        c.alignment = Alignment(horizontal="center", vertical="center")

    for r in recs:
        mt = r["merge_type"]
        # 방침대로 최종 GT를 전부 채워 둔다. 자기모순 의심 건만 확정 보류.
        ok = "" if mt in ("merged_conflict", "empty") else "Y"
        ws.append([r["sid"], mt, r["old_gt"], r["new_gt"], r["etc"],
                   r["final_gt"], ok])
        row = ws[ws.max_row]
        for c in row:
            c.alignment = Alignment(vertical="top", wrap_text=True)
        if mt in ("merged_conflict", "replaced_conflict"):
            row[1].fill = warn
        elif mt in ("merged", "replaced"):
            row[1].fill = auto
    for col, w in zip("ABCDEFG", [12, 12, 60, 60, 40, 60, 8]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx", type=Path)
    ap.add_argument("--out-dir", type=Path, default=Path("data/preprocessed"))
    ap.add_argument("--worksheet-dir", type=Path, default=Path("data/gold_sampled"))
    ap.add_argument("--show", action="store_true", help="최종 GT 전문 출력")
    ap.add_argument("--prefer-review-on-conflict", action="store_true",
                    help="기존 GT를 부정하는 문안은 병합 대신 교체 (방침 4)")
    args = ap.parse_args()

    if not args.xlsx.exists():
        sys.exit(f"파일 없음: {args.xlsx}")

    recs = extract(args.xlsx, args.prefer_review_on_conflict)
    stamp = datetime.now().strftime("%y%m%d")
    args.out_dir.mkdir(parents=True, exist_ok=True)
    jpath = args.out_dir / f"khs_review_{stamp}.json"
    jpath.write_text(json.dumps(recs, ensure_ascii=False, indent=2),
                     encoding="utf-8")

    wpath = args.worksheet_dir / "khs_gt_merge_worksheet.xlsx"
    write_worksheet(recs, wpath)

    from collections import Counter
    cnt = Counter(r["merge_type"] for r in recs)
    print(f"[import] 케이스 {len(recs)}건")
    for k in ("replaced", "replaced_conflict", "merged",
              "merged_conflict", "kept", "empty"):
        if cnt[k]:
            print(f"  {k:16s} {cnt[k]:3d}건")
    print(f"\n[import] JSON      : {jpath}")
    print(f"[import] 검토 워크시트: {wpath}")

    if args.show:
        for r in recs:
            if r["merge_type"] in ("kept", "empty"):
                continue
            print(f"\n{'=' * 72}\n[{r['sid']}] {r['merge_type']}\n{'-' * 72}")
            print(r["final_gt"])

    need = [r["sid"] for r in recs if r["merge_type"] == "merged_conflict"]
    if need:
        print(f"\n⚠ 자기모순 의심 {len(need)}건 — 확정 보류(확정?=공란): {need}")
        print("  기존 GT에 실제 처치 기록이 있는데 '특이사항 없음'이 덧붙었습니다.")
        print("  워크시트 '최종 GT'를 손보고 '확정?'에 Y 를 적으세요.")
    print(f"\n다음: python scripts/build_gold_override.py")


if __name__ == "__main__":
    main()
